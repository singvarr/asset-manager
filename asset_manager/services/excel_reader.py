from typing import TypeVar, Generic
from pathlib import Path

from openpyxl import load_workbook

T = TypeVar('T')


class ExcelReader(Generic[T]):
    """Converts contents of smart table to list of dictionaries"""
    _workbook = None

    def __init__(self, workbook_path: Path, table_name: str, remove_last_row: bool = False):
        self._workbook = ExcelReader._load_workbook(self, str(workbook_path))
        self._table_name = table_name

        self._remove_last_row = remove_last_row

    @staticmethod
    def _load_workbook(instance, workbook_path):
        if not instance._workbook:
            instance._workbook = load_workbook(workbook_path, data_only=True)

        return instance._workbook

    def run(self) -> list[T]:
        worksheet = next(
            (
                worksheet
                for worksheet in self._workbook.worksheets
                if worksheet.title == self._table_name
            ),
            None
        )

        table: list[T] = worksheet.tables.get(self._table_name)

        if not table:
            raise Exception(f'Table {self._table_name} not found')

        table_range = table.ref
        table_cells = list(worksheet[table_range])

        if self._remove_last_row:
            table_cells.pop()

        headers = [cell.value for cell in table_cells[0]]

        table_data = []

        for row in table_cells[1:]:
            row_dict = {headers[i]: cell.value for i, cell in enumerate(row)}
            table_data.append(row_dict)

        return table_data
